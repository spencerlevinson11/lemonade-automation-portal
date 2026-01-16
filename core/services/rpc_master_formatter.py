from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill


@dataclass(frozen=True)
class MasterRow:
    nld_date: Optional[dt.date]
    nld_week: Optional[int]
    customer_po: Optional[str]
    rpc_number: Optional[int]
    city: Optional[str]
    customer_name: Optional[str]
    mix_flag: str
    bucket_type: str
    quantity: int
    due_by: Optional[dt.date]
    due_week: Optional[int]


# Master sheet columns (1-indexed) based on your example
MASTER_HEADERS: List[Tuple[int, Optional[str]]] = [
    (1, "NLD"),
    (2, "Week #"),
    (3, "Departure Date"),
    (4, None),  # Customer PO# (no header in your example)
    (5, "RPC#"),
    (6, "City"),
    (7, "Customer"),
    (8, None),  # MIX (no header in your example)
    (9, "CLASSIC HQ"),
    (10, "CLASSIC"),
    (11, "NextGen HQ"),
    (12, "NextGen N2"),
    (13, "5-liter round"),
    (14, "5-liter Vase"),
    (15, "7-liter Vase"),
    (16, "7-liter Vase HQ"),
    (17, "10 Conical"),
    (18, "10 Conical NG"),
    (19, "13 Conical"),
    (20, None),
    (21, "8 liter round NG"),
    (22, "13 NextGen"),
    (23, "3 liter round"),
    (24, "MAXIMA"),
    (25, "SUB"),
    (26, "HOLD"),
    (27, "Bucket Type"),
    (28, "Due By"),
    (29, "Due week"),
    (30, "Average Transit "),
    (31, None),
]


BUCKETTYPE_TO_COLUMN: Dict[str, int] = {
    # Classic
    "10ltr Wide Classic": 10,
    "High Quality Classic": 9,
    # NextGen
    "Next Gen 10ltr": 12,
    "Next Gen 10ltr HQ": 11,
    "NG High Quality": 11,
    # Shapes
    "5ltr Round": 13,
    "5ltr Vase": 14,
    "5 ltr Vase": 14,
    "5 liter vase": 14,
    "7 ltr vase": 15,
    "7ltr Vase": 15,
    "7-ltr Vase": 15,
    "7 ltr vase HQ": 16,
    "7-ltr Vase HQ": 16,
    "10ltr Conical": 17,
    "10ltr Conical NG": 18,
    "13ltr Conical": 19,
    # NOTE: In your master sheet, the blank header column between
    # "13 Conical" and "8 liter round NG" is the standard 8 liter round.
    "8 liter round": 20,
    "8 ltr round": 20,
    "8ltr round": 20,
    "8 ltr round NG": 21,
    "8 liter round NG": 21,
    "13ltr NG": 22,
    "13ltr NextGen": 22,
    "3 liter round": 23,
    # Maxima
    "Maxima Buckets": 24,
    "Maxima Lids": 24,
    "Maxima Sets": 24,
    # Other
    "Amalia Buckets": 25,
}


# Bucket Type cell colors (matches your "bucket types and corresponding colors.xlsx").
# Any rgb value of "00000000" in the source sheet indicates no explicit fill; we skip those.
BUCKETTYPE_COLOR_RGB: Dict[str, str] = {
    "10ltr Wide Classic": "FFFCE4D6",
    "Next Gen 10ltr": "FFD9E1F2",
    "Next Gen 10ltr HQ": "FFF4B084",
    "5ltr Round": "FFFF33CC",
    "5ltr Vase": "FFD0CECE",
    "7 ltr vase HQ": "FF00B050",
    "10ltr Conical": "FFFFF2CC",
    "10ltr Conical NG": "FF9999FF",
    "13ltr Conical": "FF00B0F0",
    "8 liter round": "FF66FFCC",
    "8 ltr round": "FF66FFCC",
    "8ltr round": "FF66FFCC",
    "8 ltr round NG": "FFC65911",
    "8 liter round NG": "FFC65911",
    "13ltr NG": "FF0070C0",
    "3 liter round": "FFC6E0B4",
    "Maxima Buckets": "FFFF9999",
    "Maxima Lids": "FFFF9999",
    "Maxima Sets": "FFFF9999",
    "Amalia Buckets": "FFFF9999",
}

# Row highlight fills from your example master spreadsheet
NLD_CELL_RGB = "FFE2EFDA"      # light green used on NLD date cells
RPCNUM_CELL_RGB = "FFC6E0B4"   # light green used on RPC# cells


def _apply_cell_fill(cell, rgb: str) -> None:
    """Apply a solid fill + black font."""
    if not rgb:
        return
    cell.fill = PatternFill(patternType="solid", fgColor=rgb)
    cell.font = Font(color="FF000000")


def _apply_bucket_type_style(cell) -> None:
    """Apply the bucket-type color + readable font (matches your example master file)."""
    if not cell.value:
        return
    rgb = BUCKETTYPE_COLOR_RGB.get(str(cell.value).strip())
    if not rgb or rgb == "00000000":
        return
    cell.fill = PatternFill(patternType="solid", fgColor=rgb)
    cell.font = Font(color="FF000000")


def _safe_int(x) -> Optional[int]:
    try:
        if x is None:
            return None
        if isinstance(x, int):
            return x
        if isinstance(x, float):
            return int(x)
        s = str(x).strip()
        if not s:
            return None
        return int(float(s))
    except Exception:
        return None


def _parse_first_int(text: str) -> Optional[int]:
    if not text:
        return None
    m = re.search(r"(\d{3,})", str(text))
    return int(m.group(1)) if m else None


def _parse_month_day(text: str, year: int) -> Optional[dt.date]:
    """Parse strings like 'NLD November 10 - week 46 or asap' or 'Delivery December 22 - week 52'."""
    if not text:
        return None
    s = str(text)
    m = re.search(
        r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\b\s+(\d{1,2})",
        s,
        flags=re.IGNORECASE,
    )
    if not m:
        return None
    month_name = m.group(1).lower()
    day = int(m.group(2))
    month_map = {
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12,
    }
    month = month_map.get(month_name)
    if not month:
        return None
    return dt.date(year, month, day)


def _to_week(d: Optional[dt.date]) -> Optional[int]:
    if not d:
        return None
    return int(d.isocalendar().week)


def _align_to_friday(d: dt.date) -> dt.date:
    """Your example uses the prior Friday as NLD when the sheet says a Monday date."""
    # Monday=0 ... Sunday=6, Friday=4
    delta = (d.weekday() - 4) % 7
    return d - dt.timedelta(days=delta)


# Transit-time lookup (Average Days, Fastest Delivery)
# Provided by you; if a city is missing, output "unknown".
# Keys are normalized (lowercase, stripped).
CITY_TRANSIT_TIMES: Dict[str, Tuple[float, float]] = {
    "avondale": (47.0, 42.0),
    "belfair": (47.6, 43.0),
    "buford": (34.1, 22.0),
    "canby": (47.1, 38.0),
    "carlsbad": (50.3, 37.0),
    "carpinteria": (46.0, 40.0),
    "coppell": (36.9, 25.0),
    "dayton": (27.7, 24.0),
    "doral": (30.5, 22.0),
    "doraville": (36.6, 22.0),
    "egg harbor city": (30.0, 22.0),
    "bensenville/egv": (34.8, 24.0),
    "franklin park": (25.5, 24.0),
    "garland": (42.0, 28.0),
    "gilroy": (59.9, 39.0),
    "hyde park": (29.8, 21.0),
    "indianapolis": (36.5, 31.0),
    "itasca": (41.6, 29.0),
    "jessup": (28.5, 28.0),
    "joliet": (34.0, 25.0),
    "kansas city": (38.0, 25.0),
    "lebanon": (21.0, 21.0),
    "lombard": (37.1, 21.0),
    "long beach": (51.9, 45.0),
    "los angeles": (49.0, 42.0),
    "louisville": (39.2, 24.0),
    "mechanicsburg": (29.9, 17.0),
    "miami": (30.7, 21.0),
    "mickleton": (34.0, 27.0),
    "oxnard": (49.1, 35.0),
    "phoenix": (48.3, 38.0),
    "salt lake city": (58.0, 58.0),
    "vernon": (47.7, 35.0),
    "wilmington": (28.0, 28.0),
}


def _normalize_city_key(city: Optional[str]) -> Optional[str]:
    if not city:
        return None
    return re.sub(r"\s+", " ", str(city)).strip().lower()


def get_transit_times(city: Optional[str]) -> Tuple[object, object]:
    """Return (avg_days, fastest_days). If unknown, returns ("unknown", "unknown")."""
    key = _normalize_city_key(city)
    if not key:
        return "unknown", "unknown"
    val = CITY_TRANSIT_TIMES.get(key)
    if not val:
        return "unknown", "unknown"
    return val[0], val[1]


def _normalize_bucket_type(raw: str) -> str:
    s = (raw or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def infer_master_bucket_type(rpc_bucket_type: str) -> str:
    """Best-effort mapping from RPC 'Soort emmer' values to your master 'Bucket Type' taxonomy."""
    raw = _normalize_bucket_type(rpc_bucket_type)
    low = raw.lower()

    # "#" in your RPC sheet denotes HQ
    is_hq = ("#" in raw) or ("hq" in low)

    # Maxima
    if "maxima" in low:
        if "lid" in low:
            return "Maxima Lids"
        if "set" in low:
            return "Maxima Sets"
        return "Maxima Buckets"

    # Amalia
    if "amalia" in low:
        return "Amalia Buckets"

    # Next Gen 10
    if "next gen" in low or re.search(r"\bng\b", low):
        if "10" in low:
            return "Next Gen 10ltr HQ" if is_hq else "Next Gen 10ltr"
        if "13" in low:
            return "13ltr NG"
        if "conical" in low and "10" in low:
            return "10ltr Conical NG"

    # Conical
    if "conical" in low:
        if "13" in low:
            return "13ltr Conical"
        return "10ltr Conical NG" if ("ng" in low or "next gen" in low) else "10ltr Conical"

    # Vase
    if "vase" in low:
        if "7" in low:
            return "7 ltr vase HQ" if is_hq else "7 ltr vase"
        return "5ltr Vase"

    # Round
    if "round" in low:
        if "8" in low:
            return "8 ltr round NG" if ("ng" in low or "next gen" in low) else "8 liter round"
        if "3" in low:
            return "3 liter round"
        return "5ltr Round"

    # Classic wide 10
    if "classic" in low:
        return "High Quality Classic" if is_hq else "10ltr Wide Classic"

    # Fallback: return raw (still useful for later manual correction)
    return raw


def parse_rpc_order_xlsx(file_bytes: bytes) -> Tuple[Dict[str, object], List[MasterRow]]:
    """Parse an RPC order sheet into aggregated MasterRow entries."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # RPC number + city from the title line if present
    title = str(ws.cell(1, 1).value or "")
    rpc_number = _parse_first_int(title)

    # Customer PO#
    customer_po = None
    for r in range(1, 15):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "po" in v.lower() and "#" in v:
            customer_po = str(_parse_first_int(v) or "") or None
            break

    # Order date year (used to interpret NLD/Delivery month/day lines)
    year = dt.date.today().year
    for r in range(1, 12):
        v = ws.cell(r, 1).value
        if isinstance(v, dt.datetime):
            year = v.date().year
            break
        if isinstance(v, dt.date):
            year = v.year
            break
        if isinstance(v, str):
            # e.g. Thursday, September 04, 2025
            m = re.search(r"\b(20\d{2})\b", v)
            if m:
                year = int(m.group(1))
                break

    # NLD and Delivery rows (usually around rows 20-22 in your template)
    nld_date = None
    due_by = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if not isinstance(v, str):
            continue
        low = v.lower()
        if low.startswith("nld"):
            nld_date = _parse_month_day(v, year)
        if low.startswith("delivery"):
            due_by = _parse_month_day(v, year)
        if nld_date and due_by:
            break

    # IMPORTANT: Keep NLD exactly as shown on the RPC order (no realignment).

    nld_week = _to_week(nld_date)
    due_week = _to_week(due_by)

    # Customer name: your template places it at row ~25 col 4, but search safely
    customer_name = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if isinstance(v, str) and v.strip() and v.strip().lower() not in {"nld", "delivery"}:
            # heuristic: customer line is a standalone name, followed by address rows
            if r >= 20 and r <= 40 and ws.cell(r + 1, 4).value and ws.cell(r + 2, 4).value:
                customer_name = v.strip()
                break

    # City: try filename-ish from title, else from address line row with comma
    city = None
    m = re.search(r"#\d+\s+(.+)$", title)
    if m:
        city = m.group(1).strip()
    if not city and isinstance(ws.cell(27, 4).value, str):
        # 'Miami, FL 33122' -> Miami
        city = ws.cell(27, 4).value.split(",")[0].strip()

    # Find line-item header row
    header_row = None
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, 10)]
        if any(isinstance(x, str) and x.strip() == "Soort emmer" for x in row_vals):
            header_row = r
            break

    if not header_row:
        return (
            {
                "rpc_number": rpc_number,
                "city": city,
                "customer_name": customer_name,
                "customer_po": customer_po,
                "nld_date": nld_date,
                "due_by": due_by,
            },
            [],
        )

    # Data columns based on your template (row below header labels)
    # A: Pallets, D: Soort emmer, G: Totaal
    totals_by_type: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        soort = ws.cell(r, 4).value
        total = ws.cell(r, 7).value
        if soort is None and total is None:
            continue
        if soort is None:
            # stop once we've passed the line table
            if r > header_row + 3:
                break
            continue

        bucket_type = infer_master_bucket_type(str(soort))
        qty = _safe_int(total) or 0
        if qty <= 0:
            continue
        totals_by_type[bucket_type] = totals_by_type.get(bucket_type, 0) + qty

    rows: List[MasterRow] = []
    for bucket_type, qty in sorted(totals_by_type.items(), key=lambda x: x[0].lower()):
        rows.append(
            MasterRow(
                nld_date=nld_date,
                nld_week=nld_week,
                customer_po=customer_po,
                rpc_number=rpc_number,
                city=city,
                customer_name=customer_name,
                mix_flag="MIX",
                bucket_type=bucket_type,
                quantity=int(qty),
                due_by=due_by,
                due_week=due_week,
            )
        )

    meta = {
        "rpc_number": rpc_number,
        "city": city,
        "customer_name": customer_name,
        "customer_po": customer_po,
        "nld_date": nld_date,
        "nld_week": nld_week,
        "due_by": due_by,
        "due_week": due_week,
    }
    return meta, rows


def build_master_format_workbook(rows: Iterable[MasterRow]) -> bytes:
    """Create an .xlsx that matches your master-sheet row layout (headers at row 3, data starts at row 4)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Format"

    # leave rows 1-2 blank to match your example
    for col_idx, header in MASTER_HEADERS:
        if header is not None:
            ws.cell(3, col_idx).value = header

    # write data
    start_row = 4
    for i, r in enumerate(rows):
        rr = start_row + i

        if r.nld_date:
            nld_cell = ws.cell(rr, 1)
            nld_cell.value = dt.datetime.combine(r.nld_date, dt.time.min)
            _apply_cell_fill(nld_cell, NLD_CELL_RGB)
        ws.cell(rr, 2).value = r.nld_week
        ws.cell(rr, 4).value = _safe_int(r.customer_po) if r.customer_po else r.customer_po
        rpc_cell = ws.cell(rr, 5)
        rpc_cell.value = r.rpc_number
        _apply_cell_fill(rpc_cell, RPCNUM_CELL_RGB)
        ws.cell(rr, 6).value = r.city
        ws.cell(rr, 7).value = r.customer_name
        ws.cell(rr, 8).value = r.mix_flag

        # bucket quantity into the mapped column
        qty_col = BUCKETTYPE_TO_COLUMN.get(r.bucket_type)
        if qty_col:
            ws.cell(rr, qty_col).value = r.quantity

        # HOLD mirrors the quantity
        ws.cell(rr, 26).value = r.quantity
        bt_cell = ws.cell(rr, 27)
        bt_cell.value = r.bucket_type
        _apply_bucket_type_style(bt_cell)
        if r.due_by:
            ws.cell(rr, 28).value = dt.datetime.combine(r.due_by, dt.time.min)
        ws.cell(rr, 29).value = r.due_week

        # Transit times
        avg_days, fastest_days = get_transit_times(r.city)
        ws.cell(rr, 30).value = avg_days
        ws.cell(rr, 31).value = fastest_days

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


