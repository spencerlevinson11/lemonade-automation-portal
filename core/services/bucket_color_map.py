from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Dict, Optional

import openpyxl
from django.conf import settings


def _norm(s: str) -> str:
    return " ".join(str(s or "").strip().lower().split())


@lru_cache(maxsize=1)
def load_bucket_type_to_argb() -> Dict[str, str]:
    """Load bucket-type -> ARGB color from the bundled xlsx.

    Returns ARGB strings like 'FFC6E0B4' or 'FFFF0000'.
    """
    base_dir = Path(settings.BASE_DIR) if hasattr(settings, 'BASE_DIR') else Path(__file__).resolve().parents[2]
    path = base_dir / 'core' / 'data' / 'bucket_types_and_colors.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    mapping: Dict[str, str] = {}
    # Expect columns: bucket type | color (varies). We'll best-effort parse.
    for r in range(1, ws.max_row + 1):
        bucket = ws.cell(r, 1).value
        color = ws.cell(r, 2).value
        if not bucket or not color:
            continue
        key = _norm(bucket)
        c = str(color).strip()
        # Normalize a few possible formats.
        # If user stored like '#RRGGBB'
        if c.startswith('#') and len(c) == 7:
            rgb = c[1:]
            mapping[key] = 'FF' + rgb.upper()
            continue
        # If stored as 6-hex already
        if len(c) == 6 and all(ch in '0123456789abcdefABCDEF' for ch in c):
            mapping[key] = 'FF' + c.upper()
            continue
        # If stored as ARGB
        if len(c) == 8 and all(ch in '0123456789abcdefABCDEF' for ch in c):
            mapping[key] = c.upper()
            continue

    return mapping


def get_bucket_type_argb(bucket_type: str) -> Optional[str]:
    m = load_bucket_type_to_argb()
    return m.get(_norm(bucket_type))
