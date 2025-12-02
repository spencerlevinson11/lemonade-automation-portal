from pathlib import Path
from collections import Counter
from datetime import datetime, date

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --- CONFIGURATION ---

TEMPLATE_PATH = Path(settings.BASE_DIR) / "rpc_templates" / "RPC_template_Bensenville.xlsx"
OUTPUT_DIR = Path(settings.BASE_DIR) / "generated_rpcs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

STACKABLE_PAIRS = [
    ("Amalia White Buckets", "Amalia White Lids"),
    ("Amalia Black Buckets", "Amalia Black Lids"),
    ("Maxima Black Buckets", "Maxima Black Lids"),
    ("Maxima Buckets White", "Maxima Lids White"),
    ("Maxima Green Buckets", "Maxima Green Lids"),
]

PER_PALLET = {
    "10 Wide Standard Classic": 2800,
    "10 ltr conical Next Gen":  4050,
    "10 ltr conical black":     3960,
    "10 ltr wide NG eco":       2842,
    "13 ltr conical black":     2660,
    "13 ltr conical Next Gen":  2730,
    "5 liter vase":             6210,
    "7 liter vase #":           3240,
    "5 liter round":            3900,
    "10 liter wide classic HQ#": 2800,
    "Amalia White Buckets":     960,
    "Amalia Black Buckets":     960,
    "Amalia White Lids":        960,
    "Amalia Black Lids":        960,
    "Maxima Green Buckets":     720,
    "Maxima Buckets White":     720,
    "Maxima Black Buckets":     720,
    "Maxima Green Lids":        720,
    "Maxima Lids White":        720,
    "Maxima Black Lids":        720,
}

ARTICLE_MAP = {
    "10 ltr wide NG eco":        500103,
    "10 ltr conical black":      500103,
    "10 ltr conical Next Gen":   500107,
    "13 ltr conical Next Gen":   500131,
    "13 ltr conical black":      500130,
    "10 Wide Standard Classic":  500100,
    "10 liter classic hq":       500110,
    "5 liter vase":              500050,
    "7 liter vase #":            500071,
    "5 liter round":             500500,
    "10 liter wide classic HQ#": 500110,
    "Amalia White Buckets":      500370,
    "Amalia White Lids":         500380,
    "Amalia Black Buckets":      500371,
    "Amalia Black Lids":         500381,
    "Maxima White Buckets":      500350,
    "Maxima Buckets White":      500350,
    "Maxima Lids White":         500360,
    "Maxima Black Buckets":      500351,
    "Maxima Black Lids":         500361,
    "Maxima Green Buckets":      500390,
    "Maxima Green Lids":         500391,
}

# Aliases + normalization
ALIASES = {
    "Maxima White Buckets": "Maxima Buckets White",
    "Maxima White Lids":    "Maxima Lids White",
}


def normalize(name: str) -> str:
    n = (name or "").strip()
    return ALIASES.get(n, n)


# Map form field names -> canonical bucket names
BUCKET_FIELD_MAP = {
    "b_10_wide_standard_classic": "10 Wide Standard Classic",
    "b_10_ltr_conical_next_gen": "10 ltr conical Next Gen",
    "b_10_ltr_conical_black": "10 ltr conical black",
    "b_10_ltr_wide_ng_eco": "10 ltr wide NG eco",
    "b_13_ltr_conical_black": "13 ltr conical black",
    "b_13_ltr_conical_next_gen": "13 ltr conical Next Gen",
    "b_5_liter_vase": "5 liter vase",
    "b_7_liter_vase": "7 liter vase #",
    "b_5_liter_round": "5 liter round",
    "b_10_liter_wide_classic_hq": "10 liter wide classic HQ#",
    "b_amalia_white_buckets": "Amalia White Buckets",
    "b_amalia_black_buckets": "Amalia Black Buckets",
    "b_amalia_white_lids": "Amalia White Lids",
    "b_amalia_black_lids": "Amalia Black Lids",
    "b_maxima_green_buckets": "Maxima Green Buckets",
    "b_maxima_buckets_white": "Maxima Buckets White",
    "b_maxima_black_buckets": "Maxima Black Buckets",
    "b_maxima_green_lids": "Maxima Green Lids",
    "b_maxima_lids_white": "Maxima Lids White",
    "b_maxima_black_lids": "Maxima Black Lids",
}


# --- Core helpers (same as your original script) ---

def pair_and_leftover(buckets, pairs):
    paired, leftovers = [], []
    for bkt, lid in pairs:
        nb, nl = buckets.get(bkt, 0), buckets.get(lid, 0)
        n_pairs = min(nb, nl)
        paired.extend([(bkt, True)] * n_pairs)
        paired.extend([(lid, True)] * n_pairs)
        buckets[bkt] = nb - n_pairs
        buckets[lid] = nl - n_pairs
    leftovers = [(n, False) for n, cnt in buckets.items() for _ in range(cnt)]
    return paired, leftovers


def stack_lids_three_high(paired, leftovers, buckets, pairs):
    lids = []
    for _, lid in pairs:
        cnt = buckets.get(lid, 0)
        lids += [lid] * cnt
        buckets[lid] = 0
    if lids:
        pad = (3 - len(lids) % 3) % 3
        lids += [lids[0]] * pad
    for lid in lids:
        paired.append((lid, True))
    leftovers[:] = [(n, s) for n, s in leftovers if n not in lids]
    return paired, leftovers


def build_pallet_list(paired, leftovers, size_map):
    return [
        {"Bucket Type": n, "Pallet Size": size_map.get(n, "large"), "Stacked": s}
        for n, s in paired + leftovers
    ]


def pack_into_containers(pallets):
    # One container with everything; cap shows total items
    containers = [pallets]
    cap = len(pallets)
    return containers, cap


def format_date_info(date_val):
    """
    Format a date (datetime/date) as 'Month DD, YYYY - week WW'.
    Falls back to plain string if not a date-like value.
    """
    if isinstance(date_val, (datetime, date)):
        fmt = date_val.strftime("%B %d, %Y")
        week = date_val.isocalendar()[1]
        return f"{fmt} - week {week}"
    return str(date_val).strip() if date_val else "asap"


def write_rpc(containers, cap, po, rpc_info, nld_val, delivery_val, address_lines):
    """
    Mirrors your original write_rpc layout:

    - NLD in D25
    - Delivery in D27
    - Address block in D30–D34 (yellow box)
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    saved_files = []
    nld_text = format_date_info(nld_val)
    delivery_text = format_date_info(delivery_val)

    for idx, cont in enumerate(containers, start=1):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        sheet_name = f"RPC#{rpc_info}" if rpc_info else f"RPC{idx}"
        wb.worksheets[0].title = sheet_name
        ws["A1"].value = f"RPC Order #{rpc_info}" if rpc_info else f"RPC Order #{idx}"
        ws["A6"].value = datetime.today().strftime("%A, %B %d, %Y")
        ws["A7"].value = f"Customer PO# {po}"

        # --- NLD / Delivery exactly as original ---
        ws["D25"].value = f"NLD {nld_text} or asap"
        ws["D27"].value = f"Delivery {delivery_text}"

        # --- Clear the entire yellow address box (D30–D34) and overwrite ---
        for row in range(30, 35):
            ws[f"D{row}"].value = None

        for i, line in enumerate(address_lines):
            if i >= 5:  # max 5 lines
                break
            ws[f"D{30 + i}"].value = line
            ws[f"D{30 + i}"].alignment = Alignment(horizontal="left")

        # ---- Clear any old template totals/labels so nothing reappears ----
        for cell in ("G16", "G23", "H16", "H23"):
            ws[cell].value = None

        counts = Counter(p["Bucket Type"] for p in cont)

        # ---- Item lines ----------------------------------------------------
        row = 14
        for name, qty in counts.items():
            ws[f"A{row}"].value = qty
            ws[f"A{row}"].alignment = Alignment(horizontal="right")
            ws[f"B{row}"].value = "100 x 120"
            ws[f"C{row}"].value = PER_PALLET.get(name, "")
            ws[f"D{row}"].value = name
            ws[f"E{row}"].value = "White" if "White" in name else "Zwart"
            ws[f"F{row}"].value = ARTICLE_MAP.get(name, "")
            ws[f"G{row}"].value = qty * PER_PALLET.get(name, 0)

            for col in "ABCDEFG":
                ws[f"{col}{row}"].alignment = Alignment(
                    horizontal="right" if col == "A" else "center"
                )
            row += 1

        # ---- Summary rows (same idea as original) -------------------------
        ws["A23"].value = sum(counts.values())
        ws["A23"].alignment = Alignment(horizontal="right")
        ws["A24"].value = cap
        ws["A24"].alignment = Alignment(horizontal="right")

        total_row = row
        total_pieces_value = sum(
            qty * PER_PALLET.get(name, 0) for name, qty in counts.items()
        )

        ws[f"G{total_row}"].value = total_pieces_value
        ws[f"G{total_row}"].alignment = Alignment(horizontal="right")
        ws[f"H{total_row}"].value = "Total pieces"
        ws[f"H{total_row}"].alignment = Alignment(horizontal="left")

        # Collapse blank rows between grand total and the "pallets" area
        delete_start = total_row + 1
        if delete_start <= 22:
            ws.delete_rows(delete_start, 22 - delete_start + 1)

        filepath = OUTPUT_DIR / f"{sheet_name}.xlsx"
        wb.save(filepath)
        saved_files.append(filepath)

    return saved_files


# --- Optional Outlook draft creation (Windows only) ---

try:
    import win32com.client as win32  # type: ignore
except ImportError:
    win32 = None


def create_outlook_draft(files, subject_base, to_addrs, cc_addrs, bcc_addrs, html_body):
    """
    Try to create an Outlook draft. Returns (success: bool, message: str)
    so the caller can surface what actually happened.
    """
    if win32 is None:
        return False, "Outlook/pywin32 not available on this machine."

    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.Subject = f"New Order RPC#{subject_base}"
        mail.To = to_addrs
        mail.CC = cc_addrs
        mail.BCC = bcc_addrs
        mail.HTMLBody = html_body
        for f in files:
            mail.Attachments.Add(str(f))
        mail.Save()
        # If you want to be fancy, you can inspect mail.Parent.Name (Drafts, etc.)
        return True, "Outlook draft(s) created successfully."
    except Exception as e:
        # Bubble up the error so you can see it in the browser
        return False, f"Error creating Outlook draft: {e}"


# --- Main entry point used by Django view ---

def generate_rpc_from_form(data):
    """
    data is RpcOrderForm.cleaned_data
    Returns (files, outlook_status_message).

    files: list[Path] for generated RPC workbooks.
    outlook_status_message: string describing what happened with Outlook drafts.
    """
    po = data.get("po", "").strip()
    rpc_info = data.get("rpc_info", "").strip()
    nld = data.get("nld")
    delivery = data.get("delivery")
    company = data.get("company", "").strip()
    city_state = data.get("city_state", "").strip()
    denkers = data.get("denkers", False)

    # Build ordered address lines just like your generator sheet had:
    #   D30: company
    #   D31: street
    #   D32: city/state
    #   D33–D34: any extra lines (e.g. "Attn: Spencer ...")
    address_lines_raw = [
        company,
        (data.get("addr_line1", "") or "").strip(),
        city_state,
        (data.get("addr_line2", "") or "").strip(),
        (data.get("addr_line3", "") or "").strip(),
        (data.get("addr_line4", "") or "").strip(),
        (data.get("addr_line5", "") or "").strip(),
    ]
    address_lines = [line for line in address_lines_raw if line]

    # Build buckets dict from the bucket bank fields
    buckets = {}
    for field_name, bucket_name in BUCKET_FIELD_MAP.items():
        qty = data.get(field_name)
        if qty and qty > 0:
            key = normalize(bucket_name)
            buckets[key] = buckets.get(key, 0) + qty

    # No custom pallet-size map for now; everything defaults to "large"
    size_map = {}

    paired, leftovers = pair_and_leftover(buckets, STACKABLE_PAIRS)
    paired, leftovers = stack_lids_three_high(paired, leftovers, buckets, STACKABLE_PAIRS)
    pallets, cap = pack_into_containers(build_pallet_list(paired, leftovers, size_map))
    files = write_rpc(pallets, cap, po, rpc_info, nld, delivery, address_lines)

    # --- Outlook drafts like your original script (if possible) ---
    pickup = f"NLD {format_date_info(nld)} or asap"
    html1 = f"""
<p>Hi Annemiek and team,</p>
<p>New Order RPC#{rpc_info}</p>
<p>Can you kindly accept this order for {company} in {city_state}?</p>
<p>Please advise if the pickup date is right for you:</p>
<p><span style="background-color:yellow"><b>{pickup}</b></span></p>
<p>Kindly confirm and we will release to the forwarder.</p>
<p>Thank you!</p>

<p>Kind regards,<br>
Spencer Levinson<br>
Retriever Packaging Company LLC<br>
618 Supreme Drive<br>
Bensenville, IL, 60106<br>
708-800-6730</p>
"""
    ok1, msg1 = create_outlook_draft(
        files,
        rpc_info,
        "Annemiek.Naber@naberplastics.com;Orders@naberplastics.com",
        "jaime@retriever.pro;stan@retrieverpackaging.com",
        "spencer@retriever.pro",
        html1,
    )

    msg_parts = [msg1]

    if denkers:
        pickup2 = f'<span style="background-color:yellow"><b>{pickup}</b></span>'
        html2 = f"""
<p>Hi Paul and Esther,</p>
<p>New Order RPC#{rpc_info}</p>
<p>Please make arrangements for RPC#{rpc_info}.</p>
<p>Please communicate with Annemiek and Maud for the most convenient pickup date,</p>
<p>{pickup2}</p>
<p>Please book with the similar schedule below</p>
<p>Vessel name<br>ETD:<br>ETA:</p>
<p>When you can kindly reply with the booking confirmation.</p>
<p>Kindly confirm. Thanks!</p>

<p>Kind regards,<br>
Spencer Levinson<br>
Retriever Packaging Company LLC<br>
618 Supreme Drive<br>
Bensenville, IL, 60106<br>
708-800-6730</p>
"""
        ok2, msg2 = create_outlook_draft(
            files,
            rpc_info,
            "pv@denkersbv.nl;evdd@denkersbv.nl;digdos@denkersbv.nl",
            "stan@retrieverpackaging.com;jaime@retriever.pro",
            "spencer@retriever.pro",
            html2,
        )
        msg_parts.append(msg2)

    # Combine messages (drop empties)
    outlook_status = " / ".join(m for m in msg_parts if m)

    return files, outlook_status
