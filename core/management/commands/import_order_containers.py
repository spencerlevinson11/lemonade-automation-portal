from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.dateparse import parse_date

from core.models import Company, OrderContainer, OrderContainerLine


def _clean_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in {"nan", "none", "null"} else s


def _to_bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = _clean_str(v).lower()
    return s in {"1", "true", "t", "yes", "y", "asap"}


def _to_int(v: Any) -> Optional[int]:
    s = _clean_str(v)
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _to_date(v: Any):
    # openpyxl will return datetime/date for date cells.
    try:
        from datetime import date, datetime

        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
    except Exception:
        pass

    s = _clean_str(v)
    if not s:
        return None
    d = parse_date(s)
    return d


def _read_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [dict(r) for r in reader]


def _read_xlsx_sheet(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise CommandError(f"openpyxl is required to read xlsx files: {e}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise CommandError(f"Excel file is missing required sheet '{sheet_name}'. Found: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [(_clean_str(h) or "").strip() for h in rows[0]]
    if not any(headers):
        raise CommandError(f"Sheet '{sheet_name}' has an empty header row.")

    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        # skip fully empty rows
        if not any(x is not None and _clean_str(x) != "" for x in r):
            continue
        d: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out


class Command(BaseCommand):
    help = (
        "Import OrderContainer + OrderContainerLine records from an Excel (.xlsx) with sheets "
        "named 'containers' and 'lines', or from a CSV (containers only).\n\n"
        "Designed to work with the 'Falcon Farms Orders - Import Ready.xlsx' format we generated."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to .xlsx/.xlsm or .csv")
        parser.add_argument("--company", required=True, help="Company name to import into")
        parser.add_argument(
            "--create-company",
            action="store_true",
            help="Create the Company if it doesn't exist (recommended for first-time imports).",
        )
        parser.add_argument(
            "--mode",
            choices=["upsert", "create_only"],
            default="upsert",
            help="upsert updates existing containers matched by key fields; create_only always creates new.",
        )
        parser.add_argument("--dry-run", action="store_true", help="Validate + print actions without writing.")
        parser.add_argument(
            "--clear-lines",
            action="store_true",
            help="When importing lines: delete existing lines for matched containers before re-adding.",
        )
        parser.add_argument(
            "--match",
            default="rpc,booking,bol,po",
            help=(
                "Comma-separated match priority for upsert. Options: rpc, booking, bol, po. "
                "Default: rpc,booking,bol,po"
            ),
        )

    def _get_company(self, name: str, create_company: bool) -> Company:
        company = Company.objects.filter(name=name).first()
        if company:
            return company
        if not create_company:
            raise CommandError(
                f"Company not found: '{name}'. Re-run with --create-company to create it automatically."
            )
        return Company.objects.create(name=name)

    def _find_existing(self, company: Company, row: Dict[str, Any], match_order: List[str]) -> Optional[OrderContainer]:
        qs = OrderContainer.objects.filter(company=company)

        # These keys match the import-ready sheet columns.
        rpc = _clean_str(row.get("rpc_number"))
        booking = _clean_str(row.get("booking_number"))
        bol = _clean_str(row.get("bill_of_lading_number"))
        po_raw = _clean_str(row.get("po_number_raw"))
        po = _clean_str(row.get("po_number"))

        for key in match_order:
            if key == "rpc" and rpc:
                hit = qs.filter(rpc_number=rpc).order_by("-updated_at").first()
                if hit:
                    return hit
            if key == "booking" and booking:
                hit = qs.filter(booking_number=booking).order_by("-updated_at").first()
                if hit:
                    return hit
            if key == "bol" and bol:
                hit = qs.filter(bill_of_lading_number=bol).order_by("-updated_at").first()
                if hit:
                    return hit
            if key == "po":
                # Prefer exact raw match if present, otherwise plain po_number
                if po_raw:
                    po_num = po_raw.replace("PO#", "").strip()
                    hit = qs.filter(po_number=po_num).order_by("-updated_at").first()
                    if hit:
                        return hit
                if po:
                    hit = qs.filter(po_number=po).order_by("-updated_at").first()
                    if hit:
                        return hit
        return None

    def handle(self, *args, **opts):
        path = Path(opts["file"]).expanduser()
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        company_name = opts["company"]
        create_company = bool(opts["create_company"])
        mode = opts["mode"]
        dry_run = bool(opts["dry_run"])
        clear_lines = bool(opts["clear_lines"])

        match_raw = _clean_str(opts.get("match"))
        match_order = [m.strip().lower() for m in match_raw.split(",") if m.strip()]
        allowed = {"rpc", "booking", "bol", "po"}
        if any(m not in allowed for m in match_order):
            raise CommandError(f"Invalid --match value. Allowed: {sorted(allowed)}")
        if not match_order:
            match_order = ["rpc", "booking", "bol", "po"]

        company = self._get_company(company_name, create_company=create_company)

        containers_rows: List[Dict[str, Any]]
        lines_rows: List[Dict[str, Any]] = []

        if path.suffix.lower() == ".csv":
            containers_rows = _read_csv(path)
        else:
            containers_rows = _read_xlsx_sheet(path, "containers")
            # lines is optional
            try:
                lines_rows = _read_xlsx_sheet(path, "lines")
            except CommandError:
                lines_rows = []

        if not containers_rows:
            raise CommandError("No container rows found to import.")

        # Index lines by rpc_number and/or po_number_raw (both are present in the import-ready output)
        lines_by_rpc: Dict[str, List[Dict[str, Any]]] = {}
        lines_by_po_raw: Dict[str, List[Dict[str, Any]]] = {}
        for lr in lines_rows:
            rpc = _clean_str(lr.get("rpc_number"))
            po_raw = _clean_str(lr.get("po_number_raw"))
            if rpc:
                lines_by_rpc.setdefault(rpc, []).append(lr)
            if po_raw:
                lines_by_po_raw.setdefault(po_raw, []).append(lr)

        created = 0
        updated = 0
        created_lines = 0
        actions: List[str] = []

        with transaction.atomic():
            for row in containers_rows:
                customer_name = _clean_str(row.get("customer_name"))
                if not customer_name:
                    raise CommandError("Missing required field: customer_name")

                # Build payload
                po_number = _clean_str(row.get("po_number"))
                po_raw = _clean_str(row.get("po_number_raw"))
                if not po_number and po_raw:
                    po_number = po_raw.replace("PO#", "").strip()

                payload = {
                    "customer_name": customer_name,
                    "location_name": _clean_str(row.get("location_name")),
                    "po_number": po_number,
                    "requested_date": _to_date(row.get("requested_date")),
                    "requested_date_text": _clean_str(row.get("requested_date_text")),
                    "requested_asap": _to_bool(row.get("requested_asap")),
                    "status": _clean_str(row.get("status")),
                    "assigned_to": _clean_str(row.get("assigned_to")),
                    "rpc_number": _clean_str(row.get("rpc_number")),
                    "loading_date": _to_date(row.get("loading_date")),
                    "etd": _to_date(row.get("etd")),
                    "eta": _to_date(row.get("eta")),
                    "eta_city": _clean_str(row.get("eta_city")),
                    "estimated_delivery_date": _to_date(row.get("estimated_delivery_date")),
                    "booking_number": _clean_str(row.get("booking_number")),
                    "bill_of_lading_number": _clean_str(row.get("bill_of_lading_number")),
                    "vessel_name": _clean_str(row.get("vessel_name")),
                    "vessel_mmsi": _to_int(row.get("vessel_mmsi")),
                    "vessel_imo": _to_int(row.get("vessel_imo")),
                    "notes": _clean_str(row.get("notes")),
                }

                obj: Optional[OrderContainer] = None
                if mode == "upsert":
                    obj = self._find_existing(company, row, match_order)

                if obj is None:
                    actions.append(
                        f"CREATE rpc='{payload['rpc_number']}' po='{payload['po_number']}' loc='{payload['location_name']}'"
                    )
                    if not dry_run:
                        obj = OrderContainer.objects.create(company=company, **payload)
                    created += 1
                else:
                    actions.append(
                        f"UPDATE id={obj.id} rpc='{obj.rpc_number}' -> rpc='{payload['rpc_number']}'"
                    )
                    if not dry_run:
                        for k, v in payload.items():
                            setattr(obj, k, v)
                        obj.save()
                    updated += 1

                # Lines import (if provided)
                if lines_rows:
                    assert obj is not None
                    rpc_key = _clean_str(payload.get("rpc_number"))
                    po_key = po_raw

                    line_rows = []
                    if rpc_key and rpc_key in lines_by_rpc:
                        line_rows = lines_by_rpc[rpc_key]
                    elif po_key and po_key in lines_by_po_raw:
                        line_rows = lines_by_po_raw[po_key]

                    if line_rows:
                        if clear_lines and not dry_run:
                            obj.lines.all().delete()

                        for lr in line_rows:
                            desc = _clean_str(lr.get("item_description"))
                            pallets = _to_int(lr.get("pallets")) or 0
                            upp = _to_int(lr.get("units_per_pallet")) or 0

                            if not desc:
                                continue

                            actions.append(
                                f"  LINE container_id={obj.id} desc='{desc}' pallets={pallets} upp={upp}"
                            )
                            if not dry_run:
                                OrderContainerLine.objects.create(
                                    container=obj,
                                    item_description=desc,
                                    pallets=max(0, pallets),
                                    units_per_pallet=max(0, upp),
                                )
                            created_lines += 1

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. containers_created={created}, containers_updated={updated}, lines_created={created_lines}, dry_run={dry_run}"
            )
        )
        # Print a preview of actions to help validate.
        for a in actions[:80]:
            self.stdout.write(a)
        if len(actions) > 80:
            self.stdout.write(f"... ({len(actions) - 80} more actions)")
