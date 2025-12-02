# core/bol_generation.py
from pathlib import Path

from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook

INFO_SHEET_NAME = "BOL INFORMATION SHEET.xlsx"
TEMPLATE_NAME = "BOL Template.xlsx"

# Mapping: "information sheet row number" → list of target cells in the BOL template
# These row numbers correspond to the label rows in your BOL INFORMATION SHEET:
#  1  SHIPPERS:
#  2  CARRIER:
#  3  QUOTE #:
#  4  PRO#:
#  5  DATE:
#  6  PO#:
#  7  SEAL#:
#  8  Consignee Name
#  9  Consignee Street Address
# 10  Consignee City, State, Zip
# 11  Attn:
# 12  No of Pallets
# 13  Article Description
# 14  Specific Article
# 15  Amount of Article
# 16  Article Pallet Dimensions
# 17  Weight
# 18  Class

mapping_by_row = {
    1:  ["H2",  "H41"],   # shipper number
    2:  ["B3",  "B42"],   # carrier
    3:  ["B9",  "B48"],   # quote #
    4:  ["B10", "B49"],   # PRO #
    5:  ["H4",  "H43"],   # date
    6:  ["G10", "G49"],   # PO #
    7:  ["G11", "G50"],   # seal #
    8:  ["C12", "C51"],   # consignee name
    9:  ["C13", "C52"],   # consignee street
    10: ["C14", "C53"],   # city/state/zip
    11: ["C15", "C54"],   # attn
    12: ["A18", "A57"],   # no. pallets
    13: ["C18", "C57"],   # article description
    14: ["C19", "C58"],   # specific article
    15: ["C20", "C59"],   # amount of article
    16: ["C21", "C60"],   # pallet dims
    17: ["D18", "D57"],   # weight
    18: ["E18", "E57"],   # class
}


def _load_template_workbook():
    """
    Load the BOL template workbook from core/bol_templates.
    """
    base_dir = Path(settings.BASE_DIR)
    tpl_dir = base_dir / "core" / "bol_templates"
    template_path = tpl_dir / TEMPLATE_NAME
    wb = load_workbook(template_path)
    ws = wb.active
    return wb, ws


def _apply_mapping(values_by_row, ws):
    """
    Given a mapping of info-sheet row → value, write them into the
    correct cells in the BOL template worksheet.
    """
    for row, cells in mapping_by_row.items():
        value = values_by_row.get(row)
        if value is None:
            continue
        for cell in cells:
            ws[cell] = value


def _save_output_workbook(wb):
    """
    Save a timestamped filled-out BOL into BASE_DIR/generated_bols/.
    Returns the full path.
    """
    output_dir = Path(settings.BASE_DIR) / "generated_bols"
    output_dir.mkdir(exist_ok=True)

    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"BOL_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path


def generate_bol_from_templates():
    """
    ORIGINAL BEHAVIOR:
    Reads values from core/bol_templates/BOL INFORMATION SHEET.xlsx
    (column B of the labeled rows) and writes them into the BOL template.
    """
    base_dir = Path(settings.BASE_DIR)
    tpl_dir = base_dir / "core" / "bol_templates"
    info_path = tpl_dir / INFO_SHEET_NAME

    info_wb = load_workbook(info_path, data_only=True)
    info_ws = info_wb.active

    # 1) Read values from the info sheet (column 2 / column B)
    values_by_row = {}
    for row in mapping_by_row.keys():
        values_by_row[row] = info_ws.cell(row=row, column=2).value

    # 2) Load template and apply mapping
    wb, ws = _load_template_workbook()
    _apply_mapping(values_by_row, ws)

    # 3) Save
    return _save_output_workbook(wb)


def generate_bol_from_form(form_data: dict):
    """
    NEW BEHAVIOR:
    Accept cleaned_data from BOLForm and write directly to the BOL template.
    This bypasses the Excel info sheet and lets the web form drive the BOL.
    """

    # Map form field names → info-sheet row numbers
    field_to_row = {
        "shipper_number": 1,
        "carrier": 2,
        "quote_number": 3,
        "pro_number": 4,
        "ship_date": 5,
        "po_number": 6,
        "seal_number": 7,
        "consignee_name": 8,
        "consignee_street_address": 9,
        "consignee_city_state_zip": 10,
        "attention": 11,
        "num_pallets": 12,
        "article_description": 13,
        "specific_article": 14,
        "amount_of_article": 15,
        "pallet_dimensions": 16,
        "weight": 17,
        "freight_class": 18,
    }

    # Convert form data into the same "row → value" structure
    values_by_row = {}
    for field_name, row_num in field_to_row.items():
        value = form_data.get(field_name)
        values_by_row[row_num] = value

    wb, ws = _load_template_workbook()
    _apply_mapping(values_by_row, ws)

    return _save_output_workbook(wb)

