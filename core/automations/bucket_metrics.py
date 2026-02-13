import re
import math
from datetime import date
from typing import Dict, List
import pandas as pd

import openpyxl


import calendar

# Columns that actually represent bucket quantities (source columns)
BUCKET_COLUMNS = [
    "CLASSIC HQ",
    "CLASSIC",
    "NextGen HQ",
    "NextGen N2",
    "5-liter round",
    "5-liter Vase",
    "7-liter Vase",
    "7-liter Vase HQ",
    "10 Conical",
    "10 Conical NG",
    "13 Conical",
    "8 liter round NG",
    "13 NextGen",
    "3 liter round",
    "MAXIMA",
    "8 liter wide NIR grey",
    "10 liter wide NIR grey",
    "10 liter wideNIR grey",
    "SUB",
    "HOLD",
]

# Columns we want to show in the projection table (and in what order)
PROJECTION_COLUMNS = [
    "CLASSIC HQ",
    "CLASSIC",
    "Total Classics",
    "NextGen N2",
    "5-liter round",
    "5-liter Vase",
    "7-liter Vase",
    "7-liter Vase HQ",
    "10 Conical",
    "10 Conical NG",
    "13 Conical",
    "8 liter round NG",
    "13 NextGen",
    "3 liter round",
    "MAXIMA",
    "8 liter wide NIR grey",
    "10 liter wide NIR grey",
]

_YEAR_RE = re.compile(r"^\s*(19|20)\d{2}\s*$")


def _looks_like_year(s: str) -> bool:
    return bool(s and _YEAR_RE.match(str(s)))


def _period_to_label(p: pd.Period) -> str:
    # Sep-25 style labels
    return p.to_timestamp().strftime("%b-%y")


def _safe_key(col: str) -> str:
    # Convert a display column name into a safe HTML field key
    # Example: "CLASSIC HQ" -> "CLASSIC_HQ", "5-liter Vase" -> "5_liter_Vase"
    return col.replace("-", "_").replace(" ", "_").replace("/", "_").replace(".", "")



def _to_float(x, default: float = 0.0) -> float:
    """Convert x to float safely.

    Returns `default` for blanks/None/NaN/inf and for values that can't be cast.
    """
    try:
        if x is None:
            return default
        # pandas / numpy NaN, or blank cells that come through as NaN
        if hasattr(pd, "isna") and pd.isna(x):
            return default
        # guard infinities
        if isinstance(x, (int, float)):
            xf = float(x)
            if math.isnan(xf) or math.isinf(xf):
                return default
            return xf
        # strings, decimals, etc.
        xf = float(x)
        if math.isnan(xf) or math.isinf(xf):
            return default
        return xf
    except Exception:
        return default


def _to_int(x, default: int = 0) -> int:
    """Convert numeric x to int safely, returning default for blanks/NaN."""
    f = _to_float(x, default=float(default))
    try:
        if math.isnan(f) or math.isinf(f):
            return default
        return int(round(f))
    except Exception:
        return default



def normalize_customer_name(raw) -> str:
    """
    Normalize customer naming variants so metrics aggregate correctly.
    Returns "" for rows that should not be counted as customers (e.g., year headers).
    """
    if raw is None:
        return ""

    s = str(raw).strip()
    if not s or s.lower() in {"nan", "none"}:
        return ""

    if _looks_like_year(s):
        return ""

    s_lower = s.lower()

    # Retriever variants
    if s_lower in {
        "retriever",
        "retriever packaging",
        "retriever packaging company",
        "retriever packaging co.",
        "retriever packaging co",
        "retriever packaging corp",
        "retriever packaging corporation",
        "retriever packaging, inc",
        "retriever packaging inc",
        "retriever packaging llc",
    }:
        return "Retriever Packaging"

    # Seaside variants
    if s_lower in {"seaside", "seaside packaging"}:
        return "Seaside Packaging"

    # Mobi's variants
    if s_lower in {"mobi's", "mobis", "mobi's flowers", "mobis flowers"}:
        return "Mobi's Flowers"

    # Designers Choice variants (incl. common misspellings)
    if s_lower in {
        "designer's choice",
        "designers choice",
        "designer choice",
        "desginer's choice",
        "desginers choice",
        "desginer choice",
        "designer’s choice",  # curly apostrophe
        "desginer’s choice",  # curly apostrophe misspell
    }:
        return "Designers Choice"

    # Kendal rollups
    if s_lower in {
        "kendal",
        "kendal north",
        "kendal south",
        "kendal central",
        "kendal floral, llc",
        "kendal floral llc",
        "kendal floral",
        "kendal north bouquet co.",
        "kendal north bouquet co",
        "kendal north bouquet company",
    }:
        return "Kendal"

    # Bay State variants
    if s_lower in {
        "bay state",
        "baystate",
        "bay state & johnson's ct",
        "baystate & johnson's ct",
        "bay state and johnson's ct",
        "baystate and johnson's ct",
    }:
        return "Bay State"

    return s


def _read_master_list(uploaded_file) -> pd.DataFrame:
    """
    Reads 'Master List' and returns a cleaned dataframe with:
      date, customer, city, month (Period), plus numeric bucket columns.
    """
    xls = pd.ExcelFile(uploaded_file)
    df = pd.read_excel(xls, sheet_name="Master List")

    # In your file, row 1 (0-based index) is the true header row
    header_row = df.iloc[1]

    # Data starts at row index 3 (4th visible row)
    data = df.iloc[3:].copy()
    data.columns = header_row

    data = data.rename(columns={"NLD": "date", "Customer": "customer", "City": "city"})

    data["date"] = pd.to_datetime(data["date"], errors="coerce")
    data = data[~data["date"].isna()].copy()

    # Normalize customer names and remove invalid ones
    data["customer"] = data["customer"].apply(normalize_customer_name)
    data = data[data["customer"].astype(str).str.strip() != ""].copy()

    # Only bucket columns that exist
    bucket_cols = [c for c in BUCKET_COLUMNS if c in data.columns]

    # Ensure numeric
    data[bucket_cols] = data[bucket_cols].apply(pd.to_numeric, errors="coerce").fillna(0)

    data["month"] = data["date"].dt.to_period("M")

    return data



def _build_monthly_totals_master_list_this_year_only(uploaded_file) -> pd.DataFrame:
    """Build monthly totals from the *monthly totals rows* in the 'Master List' sheet.

    IMPORTANT: We must read THIS YEAR totals for each month ahead.
    In your Master List, the totals row is the numeric row immediately below the black separator row.
    However, the black row formatting can vary (and sometimes contains stray characters), so we do NOT
    rely on fill color. Instead, within each month block we choose the row that looks like a totals row:
      - Customer cell is blank (NOT 'Last Year Totals')
      - Many bucket columns are numeric (including zeros)
      - This prevents accidentally selecting detail rows (only 1-2 numeric cells) or last-year totals.

    We also look for the row where Customer == 'Last Year Totals' to populate (month-12) for YoY comparisons.
    """
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception:
        return pd.DataFrame()

    if "Master List" not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb["Master List"]

    # Month parsing helpers (case-insensitive)
    month_name_to_num = {calendar.month_name[i].lower(): i for i in range(1, 13)}
    month_abbr_to_num = {calendar.month_abbr[i].lower(): i for i in range(1, 13)}

    def _parse_month_num(val) -> int | None:
        if not isinstance(val, str):
            return None
        s = val.strip()
        if not s:
            return None
        # keep only letters/spaces
        s = re.sub(r"[^A-Za-z]+", " ", s).strip()
        if not s:
            return None
        token = s.split()[0].lower()
        if token in month_name_to_num:
            return month_name_to_num[token]
        if token in month_abbr_to_num:
            return month_abbr_to_num[token]
        # fuzzy prefix match
        for k, v in month_name_to_num.items():
            if k.startswith(token) or token.startswith(k[:3]):
                return v
        return None

    def _is_month_header(val) -> bool:
        """True only for pure month header labels like 'FEBRUARY' (not 'Feb-26, ...')."""
        if not isinstance(val, str):
            return False
        s = val.strip()
        if not s:
            return False
        # month header rows in your template do not contain digits or punctuation-heavy text
        if any(ch.isdigit() for ch in s):
            return False
        cleaned = re.sub(r"[^A-Za-z]", "", s).lower()
        return cleaned in month_name_to_num

    def _build_header_map(header_row: int) -> Dict[str, int]:
        hm: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip():
                hm[v.strip()] = c
        return hm

    def _cell_str(v) -> str:
        return v.strip().lower() if isinstance(v, str) else ""

    def _numeric_density(row_idx: int, header_map: Dict[str, int], bucket_cols: List[str]) -> tuple[int, float]:
        """Return (numeric_cell_count, numeric_abs_sum) across bucket columns."""
        cnt = 0
        ssum = 0.0
        for bc in bucket_cols:
            c = header_map.get(bc)
            if not c:
                continue
            v = ws.cell(row_idx, c).value
            if isinstance(v, (int, float)) and not (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
                cnt += 1
                ssum += abs(float(v))
        return cnt, ssum

    month_accum: Dict[pd.Period, Dict[str, float]] = {}
    current_year: int | None = None
    last_month_num: int | None = None

    # Find month header rows by scanning column A for pure month names
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value

        # Year marker: 2026, 2027, ...
        if isinstance(a, (int, float)) and 2000 <= int(a) <= 2100:
            current_year = int(a)
            continue

        if not _is_month_header(a) or current_year is None:
            continue

        mnum = _parse_month_num(a)
        if mnum is None:
            continue

        # Handle rollover where the year marker isn't repeated (DECEMBER -> JANUARY)
        if last_month_num is not None and mnum < last_month_num:
            current_year += 1
        last_month_num = mnum

        header_row = r - 1
        if header_row < 1:
            continue

        header_map = _build_header_map(header_row)

        # Only keep bucket columns that exist in this sheet's header row
        bucket_cols = [c for c in BUCKET_COLUMNS if c in header_map]
        if not bucket_cols:
            continue

        # Define the block end: next month header or 400 rows ahead (whichever first)
        block_end = min(ws.max_row, r + 400)
        for rr in range(r + 1, min(ws.max_row, r + 401) + 1):
            if _is_month_header(ws.cell(rr, 1).value):
                block_end = rr - 1
                break

        cust_col = header_map.get("Customer", 7)

        # Scan for this-year totals candidate rows + last-year totals row within the block
        best_totals_row = None
        best_cnt = -1
        best_sum = -1.0
        last_year_row = None

        # Threshold: totals rows have *many* numeric cells (even when some are 0/blank)
        # Use half of bucket columns, but at least 5.
        threshold = max(5, int(len(bucket_cols) * 0.5))

        for rr in range(r + 1, block_end + 1):
            cust_val = ws.cell(rr, cust_col).value
            cust_s = _cell_str(cust_val)

            if cust_s == "last year totals":
                last_year_row = rr
                continue

            # this-year totals row should have blank-ish customer cell
            if cust_s:
                continue

            cnt, ssum = _numeric_density(rr, header_map, bucket_cols)

            # Ignore detail rows (usually only 1-2 numeric cells)
            if cnt < threshold:
                continue

            # Pick the strongest totals-like row (max numeric count, then max sum)
            if cnt > best_cnt or (cnt == best_cnt and ssum > best_sum):
                best_totals_row = rr
                best_cnt = cnt
                best_sum = ssum

        if best_totals_row is None:
            continue

        period = pd.Period(f"{current_year}-{mnum:02d}", freq="M")
        month_accum.setdefault(period, {bc: 0.0 for bc in bucket_cols})

        for bc in bucket_cols:
            v = ws.cell(best_totals_row, header_map[bc]).value
            month_accum[period][bc] = _to_float(v)

        # Normalize the 10-liter wide NIR header variants into the single logical column
        if "10 liter wideNIR grey" in header_map and "10 liter wide NIR grey" in month_accum[period]:
            v_no_space = ws.cell(best_totals_row, header_map["10 liter wideNIR grey"]).value
            month_accum[period]["10 liter wide NIR grey"] = _to_float(v_no_space)

        # If we found last-year totals in this block, populate period-12
        if last_year_row is not None:
            prev_period = period - 12
            month_accum.setdefault(prev_period, {bc: 0.0 for bc in bucket_cols})
            for bc in bucket_cols:
                v = ws.cell(last_year_row, header_map[bc]).value
                month_accum[prev_period][bc] = _to_float(v)

            if "10 liter wideNIR grey" in header_map and "10 liter wide NIR grey" in month_accum[prev_period]:
                v_no_space = ws.cell(last_year_row, header_map["10 liter wideNIR grey"]).value
                month_accum[prev_period]["10 liter wide NIR grey"] = _to_float(v_no_space)

    if not month_accum:
        return pd.DataFrame()

    df = pd.DataFrame.from_dict(month_accum, orient="index").sort_index()
    df.index.name = "month"

    # Ensure all bucket columns exist
    for c in BUCKET_COLUMNS:
        if c not in df.columns:
            df[c] = 0.0

    return df

def build_monthly_totals(data: pd.DataFrame) -> pd.DataFrame:
    """
    Returns monthly totals by bucket type with PeriodIndex 'month'.
    """
    bucket_cols = [c for c in BUCKET_COLUMNS if c in data.columns]
    monthly = data.groupby("month")[bucket_cols].sum().sort_index()
    return monthly


def build_projection_table(
    monthly: pd.DataFrame,
    start_month: pd.Period,
    months_forward: int = 12,
    growth_pct_by_col: Dict[str, float] | None = None,
) -> pd.DataFrame:
    """
    Builds the projection table for months [start_month .. start_month+11]
    from existing monthly totals (which already includes future projections).

    growth_pct_by_col:
      - dict of {bucket_column_name: percent} where percent is e.g. 10 for +10%
      - applied to the forward-looking months only (not prior-year rows)
    """
    growth_pct_by_col = growth_pct_by_col or {}

    # Ensure all required columns exist in monthly (missing -> 0)
    for col in PROJECTION_COLUMNS:
        if col == "Total Classics":
            continue
        if col not in monthly.columns:
            monthly[col] = 0

    rows = []
    idx_periods: List[pd.Period] = [start_month + i for i in range(months_forward)]

    for p in idx_periods:
        if p in monthly.index:
            row = monthly.loc[p].copy()
        else:
            row = pd.Series({c: 0 for c in monthly.columns})

        out = {}
        out["Month"] = _period_to_label(p)

        classic_hq = _to_float(row.get("CLASSIC HQ", 0))
        classic = _to_float(row.get("CLASSIC", 0))

        def apply_growth(col_name: str, value: float) -> float:
            pct = _to_float(growth_pct_by_col.get(col_name, 0) or 0)
            return value * (1.0 + pct / 100.0)

        classic_hq = apply_growth("CLASSIC HQ", classic_hq)
        classic = apply_growth("CLASSIC", classic)

        out["CLASSIC HQ"] = round(classic_hq)
        out["CLASSIC"] = round(classic)
        out["Total Classics"] = round(classic_hq + classic)

        for col in PROJECTION_COLUMNS:
            if col in {"Month", "CLASSIC HQ", "CLASSIC", "Total Classics"}:
                continue
            base_val = _to_float(row.get(col, 0))
            base_val = apply_growth(col, base_val)
            out[col] = round(base_val)

        rows.append(out)

    proj_df = pd.DataFrame(rows, columns=["Month"] + PROJECTION_COLUMNS)

    end_month = start_month + (months_forward - 1)
    total_label = f"Total {start_month.year}-{str(end_month.year)[-2:]}"

    totals = {"Month": total_label}
    for col in PROJECTION_COLUMNS:
        totals[col] = _to_int(proj_df[col].sum())

    prior_periods = [p - 12 for p in idx_periods]
    prior_rows = []
    for p in prior_periods:
        if p in monthly.index:
            prior_rows.append(monthly.loc[p])
        else:
            prior_rows.append(pd.Series({c: 0 for c in monthly.columns}))

    prior_sum = pd.DataFrame(prior_rows).sum(numeric_only=True)

    prior_label = f"Total {start_month.year - 1}-{str(end_month.year - 1)[-2:]}"
    prior_totals = {"Month": prior_label}

    prior_classic_hq = _to_float(prior_sum.get("CLASSIC HQ", 0))
    prior_classic = _to_float(prior_sum.get("CLASSIC", 0))
    prior_totals["CLASSIC HQ"] = _to_int(prior_classic_hq)
    prior_totals["CLASSIC"] = _to_int(prior_classic)
    prior_totals["Total Classics"] = _to_int(prior_classic_hq + prior_classic)

    for col in PROJECTION_COLUMNS:
        if col in {"CLASSIC HQ", "CLASSIC", "Total Classics"}:
            continue
        prior_totals[col] = _to_int(prior_sum.get(col, 0))

    proj_df = pd.concat([proj_df, pd.DataFrame([totals, prior_totals])], ignore_index=True)

    return proj_df


def build_yoy_suggestions(
    monthly: pd.DataFrame,
    start_month: pd.Period,
    months_forward: int = 12,
) -> pd.DataFrame:
    """
    For each projected month and each bucket type, compute YoY delta %
    comparing month vs same month last year (month-12).
    This is used for "suggestions" in the UI.
    """
    needed_cols = [c for c in PROJECTION_COLUMNS if c != "Total Classics"]
    for col in needed_cols:
        if col not in monthly.columns:
            monthly[col] = 0

    rows = []
    for i in range(months_forward):
        p = start_month + i
        p_prev = p - 12

        cur = monthly.loc[p] if p in monthly.index else pd.Series({c: 0 for c in monthly.columns})
        prev = monthly.loc[p_prev] if p_prev in monthly.index else pd.Series({c: 0 for c in monthly.columns})

        for col in needed_cols:
            cur_val = _to_float(cur.get(col, 0))
            prev_val = _to_float(prev.get(col, 0))

            if prev_val == 0:
                yoy = None
            else:
                yoy = ((cur_val - prev_val) / prev_val) * 100.0

            rows.append(
                {
                    "Month": _period_to_label(p),
                    "Bucket Type": col,
                    "This Year (current prognosis)": _to_int(cur_val),
                    "Last Year (same month)": _to_int(prev_val),
                    "YoY %": None if yoy is None else round(yoy, 1),
                }
            )

    return pd.DataFrame(rows)


# -----------------------------
# NEW: Per-customer delta suggestions (micro-adjustments)
# Compare last year's customer-month-bucket vs this year's prognosis customer-month-bucket
# -----------------------------

def build_customer_delta_suggestions(
    data: pd.DataFrame,
    start_month: pd.Period,
    months_forward: int = 12,
) -> pd.DataFrame:
    """
    Produces rows like:
      Month | Customer | Bucket Type | Prev Year | Projection | Delta

    Where:
      Prev Year   = customer's qty in (month - 12)
      Projection  = customer's qty in (month)  [current prognosis]
      Delta       = Prev Year - Projection
    """
    # IMPORTANT: these are the bucket types you care about in projections (not SUB/HOLD/MAXIMA)
    bucket_types = [c for c in PROJECTION_COLUMNS if c != "Total Classics"]
    bucket_types = [c for c in bucket_types if c in data.columns]

    if not bucket_types:
        return pd.DataFrame(columns=["Month", "Customer", "Bucket Type", "Prev Year", "Projection", "Delta"])

    # Long table at (customer, month, bucket_type)
    long = data.melt(
        id_vars=["customer", "month"],
        value_vars=bucket_types,
        var_name="bucket_type",
        value_name="qty",
    )
    long["qty"] = pd.to_numeric(long["qty"], errors="coerce").fillna(0)

    per = (
        long.groupby(["customer", "month", "bucket_type"])["qty"]
        .sum()
        .reset_index()
    )

    # lookup: (customer, month, bucket_type) -> qty
    lookup = {
        (r["customer"], r["month"], r["bucket_type"]): _to_float(r["qty"])
        for r in per.to_dict("records")
    }

    rows = []
    for i in range(months_forward):
        p = start_month + i
        p_prev = p - 12

        customers_cur = set(per.loc[per["month"] == p, "customer"].unique())
        customers_prev = set(per.loc[per["month"] == p_prev, "customer"].unique())
        customers = customers_cur.union(customers_prev)

        for cust in sorted(customers):
            if not str(cust).strip():
                continue

            for bucket in bucket_types:
                cur_val = lookup.get((cust, p, bucket), 0.0)
                prev_val = lookup.get((cust, p_prev, bucket), 0.0)
                delta = prev_val - cur_val

                if abs(delta) < 0.5:
                    continue

                rows.append(
                    {
                        "Month": _period_to_label(p),
                        "Customer": cust,
                        "Bucket Type": bucket,
                        "Prev Year": _to_int(prev_val),
                        "Projection": _to_int(cur_val),
                        "Delta": _to_int(delta),
                    }
                )

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Month", "Customer", "Bucket Type", "Prev Year", "Projection", "Delta"])

    df["abs_delta"] = df["Delta"].abs()
    df = df.sort_values(["abs_delta", "Month", "Customer", "Bucket Type"], ascending=[False, True, True, True]).drop(
        columns=["abs_delta"]
    )

    return df


def analyze_prognosis_workbook(uploaded_file):
    """
    Existing metrics + includes:
      - monthly projection table (12 months forward from current month)
      - YoY suggestions dataframe
      - NEW: per-customer delta suggestions dataframe (micro-adjustments)
    """
    data = _read_master_list(uploaded_file)

    bucket_cols = [c for c in BUCKET_COLUMNS if c in data.columns]
    data["total_buckets"] = data[bucket_cols].sum(axis=1)

    data["month_str"] = data["month"].astype(str)

    per_customer_month = (
        data.groupby(["customer", "month_str"])["total_buckets"]
        .sum()
        .reset_index()
        .sort_values(["customer", "month_str"])
    )

    long = data.melt(
        id_vars=["customer", "city", "month_str"],
        value_vars=bucket_cols,
        var_name="bucket_type",
        value_name="qty",
    )
    long["qty"] = pd.to_numeric(long["qty"], errors="coerce").fillna(0)
    long = long[long["qty"] > 0]

    per_customer_city_item = (
        long.groupby(["customer", "city", "bucket_type"])["qty"]
        .sum()
        .reset_index()
        .sort_values(["customer", "city", "bucket_type"])
    )

    per_customer_city_item_month = (
        long.groupby(["customer", "city", "month_str", "bucket_type"])["qty"]
        .sum()
        .reset_index()
        .sort_values(["customer", "city", "month_str", "bucket_type"])
    )

    top_customers = (
        per_customer_month.groupby("customer")["total_buckets"]
        .sum()
        .reset_index()
        .sort_values("total_buckets", ascending=False)
        .head(20)
    )

    monthly = _build_monthly_totals_master_list_this_year_only(uploaded_file)
    if monthly is None or monthly.empty:
        monthly = build_monthly_totals(data)

    today = date.today()
    start_month = pd.Period(today, freq="M")

    yoy_suggestions = build_yoy_suggestions(monthly, start_month, months_forward=12)

    projection_df = build_projection_table(
        monthly=monthly,
        start_month=start_month,
        months_forward=12,
        growth_pct_by_col={},
    )

    # NEW: customer delta micro-adjustments
    customer_delta_suggestions = build_customer_delta_suggestions(
        data=data,
        start_month=start_month,
        months_forward=12,
    )

    growth_fields = []
    for col in PROJECTION_COLUMNS:
        if col == "Total Classics":
            continue
        growth_fields.append({"col": col, "key": _safe_key(col)})

    return {
        "per_customer_month": per_customer_month,
        "per_customer_city_item": per_customer_city_item,
        "per_customer_city_item_month": per_customer_city_item_month,
        "top_customers": top_customers,
        "projection_df": projection_df,
        "yoy_suggestions": yoy_suggestions,
        "customer_delta_suggestions": customer_delta_suggestions,  # NEW
        "growth_fields": growth_fields,
        "start_month_label": _period_to_label(start_month),
    }


def rebuild_projection_with_growth(uploaded_file, growth_pct_by_col: Dict[str, float]):
    """
    Recompute projections after the user enters growth % assumptions.
    (Keeps original return signature for backwards compatibility.)
    """
    data = _read_master_list(uploaded_file)
    monthly = _build_monthly_totals_master_list_this_year_only(uploaded_file)
    if monthly is None or monthly.empty:
        monthly = build_monthly_totals(data)

    today = date.today()
    start_month = pd.Period(today, freq="M")

    projection_df = build_projection_table(
        monthly=monthly,
        start_month=start_month,
        months_forward=12,
        growth_pct_by_col=growth_pct_by_col,
    )

    yoy_suggestions = build_yoy_suggestions(monthly, start_month, months_forward=12)

    return projection_df, yoy_suggestions, _period_to_label(start_month)


# OPTIONAL helper (non-breaking): use this in views if you want deltas during growth rebuild too.
def rebuild_projection_with_growth_and_customer_deltas(uploaded_file, growth_pct_by_col: Dict[str, float]):
    """
    Same as rebuild_projection_with_growth, but also returns customer_delta_suggestions.
    """
    data = _read_master_list(uploaded_file)
    monthly = _build_monthly_totals_master_list_this_year_only(uploaded_file)
    if monthly is None or monthly.empty:
        monthly = build_monthly_totals(data)

    today = date.today()
    start_month = pd.Period(today, freq="M")

    projection_df = build_projection_table(
        monthly=monthly,
        start_month=start_month,
        months_forward=12,
        growth_pct_by_col=growth_pct_by_col,
    )

    yoy_suggestions = build_yoy_suggestions(monthly, start_month, months_forward=12)

    customer_delta_suggestions = build_customer_delta_suggestions(
        data=data,
        start_month=start_month,
        months_forward=12,
    )

    return projection_df, yoy_suggestions, _period_to_label(start_month), customer_delta_suggestions









