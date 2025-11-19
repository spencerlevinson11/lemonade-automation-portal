from pathlib import Path
from django.conf import settings
from openpyxl import load_workbook

# Mapping: info-sheet row → list of template cells
mapping = {
    1:  ["H2",  "H41"],
    2:  ["B3",  "B42"],
    3:  ["B9",  "B48"],
    4:  ["B10", "B49"],
    5:  ["H4",  "H43"],
    6:  ["G10", "G49"],
    7:  ["G11", "G50"],
    8:  ["C12", "C51"],
    9:  ["C13", "C52"],
    10: ["C14", "C53"],
    11: ["C15", "C54"],
    12: ["A18", "A57"],
    13: ["C18", "C57"],
    14: ["C19", "C58"],
    15: ["C20", "C59"],
    16: ["C21", "C60"],
    17: ["D18", "D57"],
    18: ["E18", "E57"],
}


def generate_bol_from_templates() -> Path:
    """
    Server-side version of your BOL generator.

    - Reads values from 'BOL INFORMATION SHEET.xlsx'
    - Writes them into 'BOL Template.xlsx'
    - Saves a new filled-out workbook in BASE_DIR / "generated_bols"
    - Returns the Path to the generated file
    """

    base_dir = Path(settings.BASE_DIR) / "core" / "bol_templates"
    info_file = base_dir / "BOL INFORMATION SHEET.xlsx"
    template_file = base_dir / "BOL Template.xlsx"

    # 1) Read info sheet
    info_wb = load_workbook(info_file, data_only=True)
    ws_info = info_wb.active

    info_data = {
        row: ws_info.cell(row=row, column=2).value
        for row in mapping
    }

    # 2) Open template workbook
    tpl_wb = load_workbook(template_file)
    tpl_ws = tpl_wb.active

    # 3) Write values into template cells
    for row, targets in mapping.items():
        val = info_data.get(row)
        for dest in targets:
            tpl_ws[dest] = val

    # 4) Save filled-out copy
    output_dir = Path(settings.BASE_DIR) / "generated_bols"
    output_dir.mkdir(exist_ok=True)

    output_path = output_dir / "BOL_Filled_Out.xlsx"
    tpl_wb.save(output_path)

    return output_path
